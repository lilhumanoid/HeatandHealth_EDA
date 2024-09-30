import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np

# Function to calculate skewness and kurtosis
def calc_skew_kurt(x):
    return pd.Series({
        'skewness': x.skew(),
        'kurtosis': x.kurtosis()
    })

# Read and prepare data
df = pd.read_excel('WorkerHealth.xlsx')
df = df.replace('No Data', pd.NA)
df['Value'] = pd.to_numeric(df['Value'], errors='coerce')

# Basic statistics
print(df.info())
print(df.describe())

all_stats = pd.concat([df['Value'].describe(), calc_skew_kurt(df['Value'])])
print("Descriptive Statistics (including skewness and kurtosis):")
print(all_stats)

# Filter for states with all 3 years of data
state_counts = df.groupby('State')['Value'].count()
complete_states = state_counts[state_counts == 3].index
df_complete = df[df['State'].isin(complete_states)]

all_stats_complete = pd.concat([df_complete['Value'].describe(), calc_skew_kurt(df_complete['Value'])])
print("Descriptive Statistics (including skewness and kurtosis) for states with all 3 years reported:")
print(all_stats_complete)

# Visualization of distribution
plt.figure(figsize=(10, 6))
sns.histplot(df_complete['Value'].dropna(), kde=True)
plt.title('Distribution of Worker Deaths (States with All 3 Years Reported)')
plt.xlabel('Number of Deaths')
plt.ylabel('Frequency')
plt.show()

print("\nStates included in the analysis:")
print(complete_states.tolist())

# Bar plot of total deaths by state
state_totals = df_complete.groupby('State')['Value'].sum().sort_values(ascending=False)
colors = plt.cm.Greens(np.linspace(0.8, 0.2, len(state_totals)))

plt.figure(figsize=(15, 10))
bars = plt.bar(state_totals.index, state_totals.values, color=colors)
plt.title('Total Worker Deaths by State (States Reporting All 3 Years)', fontsize=16)
plt.xlabel('State', fontsize=12)
plt.ylabel('Total Number of Deaths', fontsize=12)
plt.xticks(rotation=90)

for bar in bars:
    height = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2., height, f'{height:.0f}', ha='center', va='bottom')

plt.tight_layout()
plt.show()

print(f"Number of states reporting all 3 years: {len(complete_states)}")

# Time series for Georgia
georgia_data = df[df['State'] == 'Georgia'].sort_values('Year')

plt.figure(figsize=(10, 6))
sns.lineplot(x='Year', y='Value', data=georgia_data, marker='o', linewidth=2, markersize=10)
plt.title('Worker Deaths in Georgia Over Time', fontsize=16)
plt.xlabel('Year', fontsize=12)
plt.ylabel('Number of Deaths', fontsize=12)
plt.xticks([2018, 2019, 2020])
plt.ylim(0, 100)
plt.xlim(2017.5, 2020.5)

for x, y in zip(georgia_data['Year'], georgia_data['Value']):
    if pd.notna(y):
        plt.text(x, y, f'{y:.0f}', ha='center', va='bottom', fontsize=10)

plt.tight_layout()
plt.grid(True, linestyle='--', alpha=0.7)
plt.show()

print(georgia_data[['Year', 'Value']])

# Comparison of Georgia, Texas, and California
states_to_compare = ['Georgia', 'Texas', 'California']
filtered_data = df[df['State'].isin(states_to_compare)].sort_values(['State', 'Year'])

plt.figure(figsize=(12, 8))
sns.lineplot(x='Year', y='Value', hue='State', data=filtered_data, marker='o', linewidth=2, markersize=10)
plt.title('Worker Deaths: Georgia, Texas, and California Comparison', fontsize=16)
plt.xlabel('Year', fontsize=12)
plt.ylabel('Number of Deaths', fontsize=12)
plt.xticks([2018, 2019, 2020])
y_max = filtered_data['Value'].max()
plt.ylim(0, ((y_max // 100) + 1) * 100)
plt.xlim(2017.5, 2020.5)

for state in states_to_compare:
    state_data = filtered_data[filtered_data['State'] == state]
    for x, y in zip(state_data['Year'], state_data['Value']):
        if pd.notna(y):
            plt.text(x, y, f'{y:.0f}', ha='center', va='bottom', fontsize=9)

plt.tight_layout()
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend(title='State')
plt.show()

pivot_data = filtered_data.pivot(index='Year', columns='State', values='Value')
print("\nPivot Table:")
print(pivot_data)

# Per-capita analysis
population_data = pd.read_excel('Data/StatePopulations.xlsx')

states_with_population_data = ['Alabama', 'Arizona', 'Arkansas', 'California', 'Illinois', 'Indiana', 'Kansas', 'Kentucky', 'Louisiana', 'Maryland', 'Massachusetts', 'Minnesota', 'New Jersey', 'New York', 'North Carolina', 'Ohio', 'Pennsylvania', 'South Carolina', 'Tennessee', 'Texas', 'Virginia', 'Wisconsin', 'Georgia']

death_data_filtered = df[df['State'].isin(states_with_population_data)]
population_data_filtered = population_data[population_data['State'].isin(states_with_population_data)]

merged_data = pd.merge(death_data_filtered, population_data_filtered, on=['State', 'Year'])
merged_data['DeathRate'] = (merged_data['Value'] / merged_data['Population']) * 100000

print("\nPer-capita Death Rates per 100,000 population:")
print(merged_data.pivot(index='Year', columns='State', values='DeathRate'))

average_rates = merged_data.groupby('State')['DeathRate'].mean().sort_values(ascending=False)
print("\nAverage Death Rates per 100,000 population (all years):")
print(average_rates)

highest_rate_state = average_rates.index[0]
lowest_rate_state = average_rates.index[-1]

print(f"\nState with highest average death rate: {highest_rate_state}")
print(f"State with lowest average death rate: {lowest_rate_state}")

merged_data['YearOverYearChange'] = merged_data.groupby('State')['DeathRate'].pct_change() * 100

print("\nYear-over-Year Percentage Change in Death Rates:")
print(merged_data[['State', 'Year', 'DeathRate', 'YearOverYearChange']])

# Prepare data for visualization
pivot_death_rates = merged_data.pivot(index='Year', columns='State', values='DeathRate')

# 1. Heatmap of per-capita death rates
plt.figure(figsize=(20, 10))
sns.heatmap(pivot_death_rates, annot=True, fmt='.2f', cmap='YlOrRd', cbar_kws={'label': 'Deaths per 100,000 population'})
plt.title('Heatmap of Per-Capita Worker Death Rates by State and Year', fontsize=16)
plt.ylabel('Year', fontsize=12)
plt.xlabel('State', fontsize=12)
plt.xticks(rotation=45, ha='right')
plt.tight_layout()
plt.show()

# 2. Bar plot of average death rates
plt.figure(figsize=(15, 8))
average_rates.plot(kind='bar')
plt.title('Average Per-Capita Worker Death Rates by State (2018-2020)', fontsize=16)
plt.xlabel('State', fontsize=12)
plt.ylabel('Average Deaths per 100,000 population', fontsize=12)
plt.xticks(rotation=45, ha='right')
for i, v in enumerate(average_rates):
    plt.text(i, v, f'{v:.2f}', ha='center', va='bottom')
plt.tight_layout()
plt.show()

# 3. Line plot of death rates over time for all states
# 3a Calculate average death rates
average_rates = merged_data.groupby('State')['DeathRate'].mean().sort_values(ascending=False)

# 3b Select top 3 and bottom 3 states
top_bottom_states = list(average_rates.head(3).index) + list(average_rates.tail(3).index)

# 3c Prepare data for visualization
pivot_death_rates = merged_data.pivot(index='Year', columns='State', values='DeathRate')

# 3d Create the line plot
plt.figure(figsize=(15, 10))

# 3e Plot all states with a muted color
for state in pivot_death_rates.columns:
    if state not in top_bottom_states:
        plt.plot(pivot_death_rates.index, pivot_death_rates[state], color='lightgray', linewidth=1, alpha=0.5)

# 3f Plot and label the top and bottom states with distinct colors
colors = sns.color_palette("husl", len(top_bottom_states))
for state, color in zip(top_bottom_states, colors):
    line = plt.plot(pivot_death_rates.index, pivot_death_rates[state], marker='o', linewidth=2, label=state, color=color)
    plt.text(2020, pivot_death_rates[state].iloc[-1], f' {state}', verticalalignment='center')

plt.title('Per-Capita Worker Death Rates Over Time\nHighlighting States with Highest and Lowest Average Rates', fontsize=16)
plt.xlabel('Year', fontsize=12)
plt.ylabel('Deaths per 100,000 population', fontsize=12)

# 3g Improve x-axis
plt.xticks([2018, 2019, 2020])
plt.xlim(2017.8, 2020.2)

# 3h Add a light grid
plt.grid(True, linestyle='--', alpha=0.7)

# 3i Remove the legend as we've added labels directly to the lines
plt.legend().remove()

plt.tight_layout()
plt.show()

# 3j Print out the top and bottom states for reference
print("States with highest average death rates:")
print(average_rates.head(3))
print("\nStates with lowest average death rates:")
print(average_rates.tail(3))

# Filter data for Georgia, Texas, and California
states_to_compare = ['Georgia', 'Texas', 'California']
comparison_data = merged_data[merged_data['State'].isin(states_to_compare)]

# Create a pivot table for easier plotting
pivot_comparison = comparison_data.pivot(index='Year', columns='State', values='DeathRate')

# Set up the plot
plt.figure(figsize=(12, 8))

# Plot lines for each state
colors = sns.color_palette("deep", 3)
for state, color in zip(states_to_compare, colors):
    plt.plot(pivot_comparison.index, pivot_comparison[state], marker='o', linewidth=2, label=state, color=color)

# Customize the plot
plt.title('Per-Capita Worker Death Rates: Georgia, Texas, and California', fontsize=16)
plt.xlabel('Year', fontsize=12)
plt.ylabel('Deaths per 100,000 population', fontsize=12)
plt.xticks([2018, 2019, 2020])
plt.xlim(2017.8, 2020.2)
plt.grid(True, linestyle='--', alpha=0.7)

# Add value labels
for state in states_to_compare:
    for year, rate in zip(pivot_comparison.index, pivot_comparison[state]):
        plt.text(year, rate, f'{rate:.2f}', ha='center', va='bottom', fontsize=9)

# Enhance the legend
plt.legend(title='State', loc='upper left')

# Add summary statistics
for i, state in enumerate(states_to_compare):
    avg_rate = pivot_comparison[state].mean()
    plt.text(2020.3, pivot_comparison[state].iloc[-1], f'{state}\nAvg: {avg_rate:.2f}', 
             va='center', ha='left', fontsize=10, color=colors[i])

plt.tight_layout()
plt.show()

# Print summary statistics
print("Summary Statistics:")
print(pivot_comparison.describe())

# Calculate and print percentage changes
pct_change = pivot_comparison.pct_change() * 100
print("\nPercentage Change Year-over-Year:")
print(pct_change)

# Calculate overall change from 2018 to 2020
overall_change = (pivot_comparison.loc[2020] - pivot_comparison.loc[2018]) / pivot_comparison.loc[2018] * 100
print("\nOverall Percentage Change (2018 to 2020):")
print(overall_change)

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

# Assuming merged_data is your dataframe with all the per-capita data

# Create a new Excel workbook
wb = Workbook()

# Remove the default sheet
wb.remove(wb.active)

# 1. Per-capita rates for all states and years
sheet = wb.create_sheet("Per-capita Rates")
pivot_rates = merged_data.pivot(index='Year', columns='State', values='DeathRate')
for r in dataframe_to_rows(pivot_rates, index=True, header=True):
    sheet.append(r)

# 2. Descriptive statistics for all states
sheet = wb.create_sheet("Descriptive Statistics")
desc_stats = pivot_rates.describe()
for r in dataframe_to_rows(desc_stats, index=True, header=True):
    sheet.append(r)

# 3. Year-over-year percentage change
sheet = wb.create_sheet("YoY Percentage Change")
yoy_change = pivot_rates.pct_change() * 100
for r in dataframe_to_rows(yoy_change, index=True, header=True):
    sheet.append(r)

# 4. Overall change from 2018 to 2020
sheet = wb.create_sheet("Overall Change 2018-2020")
overall_change = ((pivot_rates.loc[2020] - pivot_rates.loc[2018]) / pivot_rates.loc[2018] * 100).to_frame("Percent Change")
for r in dataframe_to_rows(overall_change, index=True, header=True):
    sheet.append(r)

# 5. Summary stats for each state
for state in pivot_rates.columns:
    sheet = wb.create_sheet(f"{state} Summary")
    state_data = merged_data[merged_data['State'] == state]
    summary = pd.DataFrame({
        'Year': state_data['Year'],
        'Death Rate': state_data['DeathRate'],
        'Actual Deaths': state_data['Value'],
        'Population': state_data['Population']
    })
    for r in dataframe_to_rows(summary, index=False, header=True):
        sheet.append(r)
    
    # Add some summary statistics
    sheet.append([])
    sheet.append(['Average Death Rate', state_data['DeathRate'].mean()])
    sheet.append(['Min Death Rate', state_data['DeathRate'].min()])
    sheet.append(['Max Death Rate', state_data['DeathRate'].max()])
    sheet.append(['Total Deaths', state_data['Value'].sum()])

# Basic styling
for sheet in wb:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
wb.save("Per_Capita_Worker_Deaths_Analysis.xlsx")

print("Excel file 'Per_Capita_Worker_Deaths_Analysis.xlsx' has been created with all the analyses.")